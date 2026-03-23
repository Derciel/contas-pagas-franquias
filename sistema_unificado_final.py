import streamlit as st
import pandas as pd
import numpy as np
import plotly.graph_objects as go
import plotly.express as px
import os
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
import warnings
warnings.filterwarnings('ignore')

# Configuração da página
st.set_page_config(
    page_title="Sistema de Análise de Franquias",
    page_icon="🚀",
    layout="wide",
    initial_sidebar_state="expanded"
)

# Configuração de estilo
st.markdown("""
<style>
    .main-header {
        background: linear-gradient(90deg, #667eea 0%, #764ba2 100%);
        padding: 2rem;
        border-radius: 10px;
        margin-bottom: 2rem;
        text-align: center;
        color: white;
    }
    
    .kpi-section {
        background: #f8f9fa;
        padding: 1rem;
        border-radius: 10px;
        margin: 1rem 0;
    }
    
    .kpi-title {
        font-size: 1.2rem;
        font-weight: bold;
        color: #495057;
        margin-bottom: 1rem;
    }
    
    .metric-card {
        background: white;
        padding: 1.5rem;
        border-radius: 10px;
        box-shadow: 0 4px 6px rgba(0,0,0,0.1);
        text-align: center;
        transition: transform 0.2s;
        border-left: 4px solid #667eea;
    }
    
    .metric-card:hover {
        transform: translateY(-2px);
    }
    
    .metric-card .icon {
        font-size: 2rem;
        margin-bottom: 0.5rem;
    }
    
    .metric-card h3 {
        color: #6c757d;
        font-size: 0.9rem;
        margin: 0;
        font-weight: normal;
    }
    
    .metric-card h2 {
        color: #212529;
        font-size: 1.8rem;
        margin: 0.5rem 0;
        font-weight: bold;
    }
    
    .metric-card small {
        color: #6c757d;
        font-size: 0.8rem;
    }
    
    .ativo-card {
        border-left-color: #28a745;
    }
    
    .inativo-card {
        border-left-color: #dc3545;
    }
    
    .chart-container {
        background: white;
        padding: 1rem;
        border-radius: 10px;
        box-shadow: 0 2px 4px rgba(0,0,0,0.1);
        margin: 1rem 0;
    }
    
    .ai-command-box {
        background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
        padding: 1.5rem;
        border-radius: 10px;
        color: white;
    }
</style>
""", unsafe_allow_html=True)

class SistemaUnificadoFinal:
    def __init__(self):
        self.faturamento_data = None
        self.contas_receber_data = None
        self.clientes_data = None
        self.franquias_status = None
        self.data_cutoff = pd.Timestamp('2025-11-01')
    
    def carregar_faturamento(self, file_path):
        """Carrega dados de faturamento"""
        print("📊 Carregando dados de faturamento...")
        
        try:
            self.faturamento_data = pd.read_excel(file_path)
            print(f"✅ Faturamento carregado: {len(self.faturamento_data)} registros")
            return True
        except Exception as e:
            print(f"❌ Erro ao carregar faturamento: {str(e)}")
            return False
    
    def carregar_contas_receber(self, file_path):
        """Carrega dados de contas a receber"""
        print("💰 Carregando dados de contas a receber...")
        
        try:
            self.contas_receber_data = pd.read_excel(file_path)
            print(f"✅ Contas a receber carregadas: {len(self.contas_receber_data)} registros")
            return True
        except Exception as e:
            print(f"❌ Erro ao carregar contas a receber: {str(e)}")
            return False
    
    def carregar_clientes(self, file_path):
        """Carrega dados de clientes"""
        print("👥 Carregando dados de clientes...")
        
        try:
            self.clientes_data = pd.read_excel(file_path)
            print(f"✅ Clientes carregados: {len(self.clientes_data)} registros")
            return True
        except Exception as e:
            print(f"❌ Erro ao carregar clientes: {str(e)}")
            return False
    
    def analisar_status_franquias(self):
        """Analisa status das franquias (ativas/inativas)"""
        print("🏢 Analisando status das franquias...")
        
        if self.faturamento_data is None:
            return None
        
        df = self.faturamento_data.copy()
        
        # Converter data de emissão
        if 'Data Emissao' in df.columns:
            df['Data Emissao'] = pd.to_datetime(df['Data Emissao'], errors='coerce')
        
        # Classificar clientes por última compra
        ultima_compra = df.groupby('Cliente ID')['Data Emissao'].max().reset_index()
        ultima_compra['Status'] = ultima_compra['Data Emissao'].apply(
            lambda x: 'Ativo' if pd.notna(x) and x >= self.data_cutoff else 'Inativo'
        )
        
        # Mesclar com dados de franquia
        if 'Franquias' in df.columns:
            franquia_info = df[['Cliente ID', 'Cliente Faturamento', 'Franquias']].drop_duplicates()
            self.franquias_status = pd.merge(ultima_compra, franquia_info, on='Cliente ID', how='left')
        else:
            self.franquias_status = ultima_compra
        
        print(f"✅ Análise de franquias concluída:")
        print(f"   🟢 Franquias ATIVAS: {len(self.franquias_status[self.franquias_status['Status'] == 'Ativo'])}")
        print(f"   🔴 Franquias INATIVAS: {len(self.franquias_status[self.franquias_status['Status'] == 'Inativo'])}")
        print(f"   Total de Franquias: {len(self.franquias_status)}")
        
        return self.franquias_status
    
    def analisar_por_franquia(self):
        """Analisa dados por franquia"""
        if self.faturamento_data is None or self.franquias_status is None:
            return None, None
        
        df = self.faturamento_data.copy()
        
        # Converter data
        if 'Data Emissao' in df.columns:
            df['Data Emissao'] = pd.to_datetime(df['Data Emissao'], errors='coerce')
        
        # Mesclar com status
        df_com_status = pd.merge(df, self.franquias_status[['Cliente ID', 'Status']], on='Cliente ID', how='left')
        
        # Análise por franquia
        if 'Franquias' in df_com_status.columns:
            pivot_franquia = df_com_status.pivot_table(
                index='Franquias',
                columns='Status',
                values='R$ Total',
                aggfunc='sum',
                fill_value=0
            ).reset_index()
            
            # Adicionar colunas se não existirem
            if 'Ativo' not in pivot_franquia.columns:
                pivot_franquia['Ativo'] = 0
            if 'Inativo' not in pivot_franquia.columns:
                pivot_franquia['Inativo'] = 0
            
            pivot_franquia['Total Faturado'] = pivot_franquia['Ativo'] + pivot_franquia['Inativo']
            pivot_franquia['Taxa Inatividade'] = (pivot_franquia['Inativo'] / (pivot_franquia['Ativo'] + pivot_franquia['Inativo']) * 100).round(2)
            
            # Detalhe por franquia
            resumo_franquia = df_com_status.groupby('Franquias').agg({
                'Cliente ID': 'nunique',
                'R$ Total': 'sum',
                'Data Emissao': ['min', 'max']
            }).round(2)
            
            # Achatar colunas
            resumo_franquia.columns = ['Clientes Únicos', 'Total Faturado', 'Primeira Compra', 'Última Compra']
            resumo_franquia = resumo_franquia.reset_index()
        else:
            pivot_franquia = pd.DataFrame()
            resumo_franquia = pd.DataFrame()
        
        # Ordenar por total faturado
        if not pivot_franquia.empty:
            pivot_franquia = pivot_franquia.sort_values('Total Faturado', ascending=False)
        
        print(f"✅ Análise por franquia concluída:")
        print(f"   🏢 Total de franquias: {len(pivot_franquia)}")
        
        return pivot_franquia, resumo_franquia
    
    def analisar_contas_receber(self):
        """Analisa dados de contas a receber com correção de vendedores e dados de pagamentos"""
        print("💰 Analisando contas a receber...")
        
        if self.contas_receber_data is None:
            return None
        
        df_contas = self.contas_receber_data.copy()
        
        # Detectar coluna de valor
        coluna_valor = None
        colunas_valor_possiveis = ['R$ Valor', 'Valor', 'R$ Total', 'Total', 'Valor Total', 'Vlr Total', 'Vl. Total']
        
        for col in colunas_valor_possiveis:
            if col in df_contas.columns:
                coluna_valor = col
                break
        
        # Detectar coluna de valor pago
        coluna_pago = None
        colunas_pago_possiveis = ['R$ Total Pago', 'Valor Pago', 'Total Pago', 'Pago']
        
        for col in colunas_pago_possiveis:
            if col in df_contas.columns:
                coluna_pago = col
                break
        
        # Estatísticas básicas com precisão
        total_contas = len(df_contas)
        valor_total = 0
        valor_pago = 0
        
        if coluna_valor:
            valor_total = df_contas[coluna_valor].sum()
        
        if coluna_pago:
            valor_pago = df_contas[coluna_pago].sum()
        
        # Calcular valor pendente com precisão
        valor_pendente = valor_total - valor_pago
        
        # Status por vencimento
        status_vencimento = pd.Series()
        
        # Detectar coluna de data de vencimento
        coluna_vencimento = None
        colunas_vencimento_possiveis = ['Data Vencimento', 'Vencimento', 'Dt Vencimento', 'Data Vencto']
        
        for col in colunas_vencimento_possiveis:
            if col in df_contas.columns:
                coluna_vencimento = col
                break
        
        if coluna_vencimento:
            # Converter data se necessário
            if df_contas[coluna_vencimento].dtype == 'object':
                df_contas[coluna_vencimento] = pd.to_datetime(df_contas[coluna_vencimento], errors='coerce')
            
            hoje = pd.Timestamp.now()
            df_contas['Status Vencimento'] = df_contas[coluna_vencimento].apply(
                lambda x: 'Vencida' if pd.notna(x) and x < hoje else 'A Vencer' if pd.notna(x) else 'Sem Data'
            )
            
            status_vencimento = df_contas['Status Vencimento'].value_counts()
        
        # Análise por status de pagamento
        status_pagamento = pd.Series()
        if 'Status' in df_contas.columns:
            status_pagamento = df_contas['Status'].value_counts()
        
        # CORREÇÃO PRECISA: Usar vendedores corrigidos dos clientes
        contas_por_vendedor = pd.DataFrame()
        
        # Detectar coluna de vendedor corrigido
        coluna_vendedor = None
        colunas_vendedor_possiveis = ['Vendedor_corrigido', 'Vendedor', 'Vendedor_cliente']
        
        for col in colunas_vendedor_possiveis:
            if col in df_contas.columns:
                coluna_vendedor = col
                break
        
        if coluna_vendedor:
            # Análise por vendedor corrigido com dados de pagamento PRECISOS
            agg_dict = {
                coluna_valor: 'sum',
                'Status Vencimento': lambda x: (x == 'Vencida').sum(),
            }
            
            if coluna_pago:
                agg_dict[coluna_pago] = 'sum'
            
            contas_por_vendedor = df_contas.groupby(coluna_vendedor).agg(agg_dict).round(2)
            
            # Adicionar quantidade de contas
            contas_por_vendedor['Quantidade'] = df_contas.groupby(coluna_vendedor).size()
            
            # Calcular valor pendente por vendedor com precisão
            if coluna_pago:
                contas_por_vendedor['Valor Pendente'] = contas_por_vendedor[coluna_valor] - contas_por_vendedor[coluna_pago]
                # Calcular percentual pago com precisão
                contas_por_vendedor['% Pago'] = round(contas_por_vendedor[coluna_pago] / contas_por_vendedor[coluna_valor] * 100, 2)
                # Achatar colunas
                contas_por_vendedor.columns = ['Valor Total', 'Contas Vencidas', 'Valor Pago', 'Valor Pendente', 'Quantidade', '% Pago']
            else:
                contas_por_vendedor.columns = ['Valor Total', 'Contas Vencidas']
                contas_por_vendedor['Valor Pago'] = 0
                contas_por_vendedor['Valor Pendente'] = contas_por_vendedor['Valor Total']
                contas_por_vendedor['% Pago'] = 0
                # Achatar colunas
                contas_por_vendedor.columns = ['Valor Total', 'Contas Vencidas', 'Valor Pago', 'Valor Pendente', 'Quantidade', '% Pago']
            
            # Calcular percentual de vencidas
            contas_por_vendedor['% Vencidas'] = round(contas_por_vendedor['Contas Vencidas'] / contas_por_vendedor['Quantidade'] * 100, 2)
            
            # Reordenar colunas para melhor visualização
            if coluna_pago:
                contas_por_vendedor = contas_por_vendedor[['Valor Total', 'Valor Pago', 'Valor Pendente', 'Quantidade', 'Contas Vencidas', '% Pago', '% Vencidas']]
            else:
                contas_por_vendedor = contas_por_vendedor[['Valor Total', 'Valor Pago', 'Valor Pendente', 'Quantidade', 'Contas Vencidas', '% Pago', '% Vencidas']]
            
            # Ordenar por valor total
            contas_por_vendedor = contas_por_vendedor.sort_values('Valor Total', ascending=False)
        
        # Análise por cliente com dados precisos
        contas_por_cliente = pd.DataFrame()
        
        # Detectar coluna de cliente
        coluna_cliente = None
        colunas_cliente_possiveis = ['Nome Fantasia', 'Cliente', 'Nome Cliente', 'Cliente Nome', 'Razão Social_cliente', 'Razão Social', 'Nome']
        
        for col in colunas_cliente_possiveis:
            if col in df_contas.columns:
                coluna_cliente = col
                break
        
        if coluna_cliente:
            agg_dict_cliente = {
                coluna_valor: 'sum',
                'Status Vencimento': 'count',
            }
            
            if coluna_pago:
                agg_dict_cliente[coluna_pago] = 'sum'
            
            contas_por_cliente = df_contas.groupby(coluna_cliente).agg(agg_dict_cliente).reset_index()
            
            if coluna_pago:
                contas_por_cliente.columns = ['Cliente', 'Valor Total', 'Quantidade', 'Valor Pago']
                contas_por_cliente['Valor Pendente'] = contas_por_cliente['Valor Total'] - contas_por_cliente['Valor Pago']
            else:
                contas_por_cliente.columns = ['Cliente', 'Valor Total', 'Quantidade']
                contas_por_cliente['Valor Pago'] = 0
                contas_por_cliente['Valor Pendente'] = contas_por_cliente['Valor Total']
            
            # Ordenar por valor total
            contas_por_cliente = contas_por_cliente.sort_values('Valor Total', ascending=False)
        
        # Cálculos precisos de status
        if coluna_pago:
            # Calcular contas pagas com precisão
            contas_pagas = len(df_contas[df_contas[coluna_pago] > 0])
            contas_abertas = total_contas - contas_pagas
            taxa_pagamento = round(contas_pagas / total_contas * 100, 2) if total_contas > 0 else 0
        else:
            contas_pagas = 0
            contas_abertas = total_contas
            taxa_pagamento = 0
        
        print(f"✅ Análise de contas a receber concluída:")
        print(f"   📄 Total de contas: {total_contas}")
        print(f"   💰 Valor total: R$ {valor_total:,.2f}")
        print(f"   💵 Valor pago: R$ {valor_pago:,.2f}")
        print(f"   ⏳ Valor pendente: R$ {valor_pendente:,.2f}")
        print(f"   ✅ Contas pagas: {contas_pagas} ({taxa_pagamento}%)")
        print(f"   ⏳ Contas abertas: {contas_abertas} ({100-taxa_pagamento}%)")
        print(f"   📋 Coluna valor utilizada: {coluna_valor}")
        print(f"   📋 Coluna pago utilizada: {coluna_pago}")
        print(f"   📋 Coluna vencimento utilizada: {coluna_vencimento}")
        print(f"   📋 Coluna cliente utilizada: {coluna_cliente}")
        print(f"   📋 Coluna vendedor utilizada: {coluna_vendedor}")
        print(f"   🎯 Total de vendedores: {len(contas_por_vendedor) if not contas_por_vendedor.empty else 0}")
        
        return {
            'df_contas': df_contas,
            'total_contas': total_contas,
            'valor_total': valor_total,
            'valor_pago': valor_pago,
            'valor_pendente': valor_pendente,
            'contas_pagas': contas_pagas,
            'contas_abertas': contas_abertas,
            'taxa_pagamento': taxa_pagamento,
            'status_vencimento': status_vencimento,
            'status_pagamento': status_pagamento,
            'contas_por_cliente': contas_por_cliente,
            'contas_por_vendedor': contas_por_vendedor,
            'coluna_valor': coluna_valor,
            'coluna_pago': coluna_pago,
            'coluna_vencimento': coluna_vencimento,
            'coluna_cliente': coluna_cliente,
            'coluna_vendedor': coluna_vendedor
        }
    
    def gerar_relatorio_completo(self, nome_arquivo_saida='sistema_unificado_final.xlsx'):
        """Gera relatório Excel completo com todos os sistemas"""
        print(f"📝 Gerando relatório completo: {nome_arquivo_saida}")
        
        # Criar writer Excel
        with pd.ExcelWriter(nome_arquivo_saida, engine='openpyxl') as writer:
            
            # Aba 1: Resumo Geral
            self._criar_aba_resumo_geral(writer)
            
            # Aba 2: Relatório 1 - Franquias ATIVAS/INATIVAS
            if self.franquias_status is not None:
                pivot_franquia, resumo_franquia = self.analisar_por_franquia()
                if pivot_franquia is not None and not pivot_franquia.empty:
                    pivot_franquia.to_excel(writer, sheet_name='Franquias Resumo', index=True)
                if not resumo_franquia.empty:
                    resumo_franquia.to_excel(writer, sheet_name='Franquias Detalhe', index=False)
                self.franquias_status.to_excel(writer, sheet_name='Franquias Status', index=False)
                
                # NOVO: Criar abas separadas por franquia
                self._criar_abas_franquias_separadas(writer)
            
            # Aba 3: Relatório 2 - Contas a Receber
            if self.contas_receber_data is not None:
                analise_contas = self.analisar_contas_receber()
                if analise_contas is not None:
                    analise_contas['df_contas'].to_excel(writer, sheet_name='Contas a Receber', index=False)
                    
                    # Resumo de contas a receber COM DADOS PRECISOS
                    resumo_contas = pd.DataFrame({
                        'Métrica': ['Total de Contas', 'Valor Total', 'Valor Pago', 'Valor Pendente', 'Contas Pagas', 'Contas Abertas', 'Taxa Pagamento', 'Contas Vencidas', 'Contas a Vencer'],
                        'Valor': [
                            analise_contas['total_contas'],
                            f"R$ {analise_contas['valor_total']:,.2f}",
                            f"R$ {analise_contas['valor_pago']:,.2f}",
                            f"R$ {analise_contas['valor_pendente']:,.2f}",
                            analise_contas['contas_pagas'],
                            analise_contas['contas_abertas'],
                            f"{analise_contas['taxa_pagamento']:.1f}%",
                            analise_contas['status_vencimento'].get('Vencida', 0),
                            analise_contas['status_vencimento'].get('A Vencer', 0)
                        ]
                    })
                    resumo_contas.to_excel(writer, sheet_name='Contas Resumo', index=False)
                    
                    # Análise por vendedor corrigido
                    if not analise_contas['contas_por_vendedor'].empty:
                        analise_contas['contas_por_vendedor'].to_excel(writer, sheet_name='Contas por Vendedor', index=False)
            
            # Aba 4: Dados de Clientes
            if self.clientes_data is not None:
                self.clientes_data.to_excel(writer, sheet_name='Dados Clientes', index=False)
            
            # Aba 5: Dados Brutos
            if self.faturamento_data is not None:
                self.faturamento_data.to_excel(writer, sheet_name='Dados Faturamento', index=False)
            
            if self.contas_receber_data is not None:
                self.contas_receber_data.to_excel(writer, sheet_name='Dados Contas Receber', index=False)
            
            # Formatar planilha
            self._formatar_planilha(writer)
        
        print(f"✅ Relatório completo gerado: {nome_arquivo_saida}")
        return nome_arquivo_saida
    
    def _criar_aba_resumo_geral(self, writer):
        """Cria aba de resumo geral"""
        resumo_data = []
        
        # Resumo Franquias
        if self.franquias_status is not None:
            ativas = len(self.franquias_status[self.franquias_status['Status'] == 'Ativo'])
            inativas = len(self.franquias_status[self.franquias_status['Status'] == 'Inativo'])
            total_franquias = len(self.franquias_status)
            
            resumo_data.extend([
                ['RELATÓRIO 1 - FRANQUIAS', '', '', ''],
                ['Total de Franquias', total_franquias, '', ''],
                ['Franquias ATIVAS', ativas, f'{(ativas/total_franquias*100):.1f}%', ''],
                ['Franquias INATIVAS', inativas, f'{(inativas/total_franquias*100):.1f}%', ''],
                ['', '', '', ''],
            ])
            
            if self.faturamento_data is not None:
                faturamento_total = self.faturamento_data['R$ Total'].sum()
                resumo_data.extend([
                    ['Faturamento Total', f"R$ {faturamento_total:,.2f}", '', ''],
                    ['', '', '', ''],
                ])
        
        # Resumo Contas a Receber
        if self.contas_receber_data is not None:
            analise_contas = self.analisar_contas_receber()
            if analise_contas is not None:
                resumo_data.extend([
                    ['RELATÓRIO 2 - CONTAS A RECEBER', '', '', ''],
                    ['Total de Contas', analise_contas['total_contas'], '', ''],
                    ['Valor Total', f"R$ {analise_contas['valor_total']:,.2f}", '', ''],
                    ['Valor Pago', f"R$ {analise_contas['valor_pago']:,.2f}", '', ''],
                    ['Valor Pendente', f"R$ {analise_contas['valor_pendente']:,.2f}", '', ''],
                    ['Contas Vencidas', analise_contas['status_vencimento'].get('Vencida', 0), '', ''],
                    ['Contas a Vencer', analise_contas['status_vencimento'].get('A Vencer', 0), '', ''],
                    ['', '', '', ''],
                ])
        
        # Criar DataFrame
        resumo_df = pd.DataFrame(resumo_data, columns=['Métrica', 'Valor', 'Percentual', 'Observações'])
        resumo_df.to_excel(writer, sheet_name='RESUMO GERAL', index=False)
    
    def _criar_abas_franquias_separadas(self, writer):
        """Cria abas separadas para cada franquia com dados detalhados"""
        print("📊 Criando abas separadas por franquia...")
        
        if self.faturamento_data is None or self.franquias_status is None:
            return
        
        # Obter lista de franquias únicas
        if 'Franquias' in self.faturamento_data.columns:
            franquias_unicas = self.faturamento_data['Franquias'].dropna().unique()
            
            # Limitar para não criar muitas abas (máximo 20 franquias)
            if len(franquias_unicas) > 20:
                franquias_unicas = franquias_unicas[:20]
                print(f"   ⚠️ Limitando para as 20 principais franquias")
            
            for franquia in franquias_unicas:
                # Filtrar dados da franquia
                dados_franquia = self.faturamento_data[self.faturamento_data['Franquias'] == franquia].copy()
                
                # Adicionar status do cliente com DEPARA CORRETA
                if self.franquias_status is not None:
                    status_map = dict(zip(self.franquias_status['Cliente ID'], self.franquias_status['Status']))
                    dados_franquia['Status Cliente'] = dados_franquia['Cliente ID'].map(status_map)
                    
                    # NOVO: Adicionar vendedor corrigido do cliente
                    if self.clientes_data is not None:
                        # Criar mapa de vendedores corrigidos dos clientes
                        vendedor_map = dict(zip(self.clientes_data['Cliente ID'], self.clientes_data['Vendedor_corrigido']))
                        
                        # Adicionar vendedor corrigido à franquia
                        dados_franquia['Vendedor Corrigido'] = dados_franquia['Cliente ID'].map(vendedor_map)
                        
                        # NOVO: Adicionar dados do cliente DGA
                        # Verificar colunas disponíveis no clientes_data
                        colunas_cliente = []
                        for col in ['Nome Fantasia', 'Razão Social', 'CNPJ', 'Telefone', 'Email']:
                            if col in self.clientes_data.columns:
                                colunas_cliente.append(col)
                        
                        if colunas_cliente:
                            cliente_map = dict(zip(self.clientes_data['Cliente ID'], 
                                                                self.clientes_data[colunas_cliente].to_dict('records')))
                            
                            # Adicionar informações do cliente DGA
                            for col in colunas_cliente:
                                dados_franquia[f'Cliente {col}'] = dados_franquia['Cliente ID'].apply(
                                    lambda x: cliente_map.get(x, {}).get(col, 'N/A')
                                )
                
                # Nome válido para aba (máximo 31 caracteres)
                nome_aba = str(franquia)[:31].replace('/', '-').replace('\\', '-').replace('*', '-').replace('?', '-').replace(':', '-').replace('[', '-').replace(']', '-')
                
                # Se o nome for muito curto ou vazio, usar um padrão
                if len(nome_aba.strip()) < 3:
                    nome_aba = f"Franquia_{len(franquias_unicas) - list(franquias_unicas).index(franquia)}"
                
                try:
                    # Criar aba para a franquia
                    dados_franquia.to_excel(writer, sheet_name=nome_aba, index=False)
                    print(f"   ✅ Aba criada: {nome_aba} ({len(dados_franquia)} registros)")
                except Exception as e:
                    print(f"   ❌ Erro ao criar aba {nome_aba}: {str(e)}")
                    # Tentar com nome simplificado
                    try:
                        nome_simplificado = f"F_{len(franquias_unicas) - list(franquias_unicas).index(franquia)}"
                        dados_franquia.to_excel(writer, sheet_name=nome_simplificado, index=False)
                        print(f"   ✅ Aba criada com nome simplificado: {nome_simplificado}")
                    except:
                        print(f"   ❌ Não foi possível criar aba para a franquia: {franquia}")
            
            print(f"✅ {len(franquias_unicas)} abas de franquias criadas")
        else:
            print("❌ Coluna 'Franquias' não encontrada nos dados")
    
    def _formatar_planilha(self, writer):
        """Formata a planilha Excel"""
        try:
            workbook = writer.book
            
            # Formatos
            header_font = Font(bold=True, color='FFFFFF')
            header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
            alignment = Alignment(horizontal='center', vertical='center')
            
            for sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
                
                # Formatar cabeçalho
                for cell in sheet[1]:
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = alignment
                
                # Ajustar largura das colunas
                for column in sheet.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    
                    adjusted_width = min(max_length + 2, 50)
                    sheet.column_dimensions[column_letter].width = adjusted_width
        except Exception as e:
            print(f"Aviso: Erro na formatação da planilha: {str(e)}")

def mostrar_dashboard(sistema):
    """Mostra o dashboard principal com KPIs e gráficos"""
    
    st.markdown("""
    <div class="main-header">
        <h1>🚀 Sistema de Análise de Franquias</h1>
        <p>Análise Completa de Franquias ATIVAS/INATIVAS e Contas Pagas por Vendedor</p>
    </div>
    """, unsafe_allow_html=True)
    
    # Verificar dados carregados
    tem_faturamento = sistema.faturamento_data is not None
    tem_contas = sistema.contas_receber_data is not None
    
    if not tem_faturamento and not tem_contas:
        st.warning("⚠️ Nenhum dado carregado. Por favor, faça upload dos arquivos na sidebar.")
        return
    
    # KPIs para Franquias
    if tem_faturamento:
        sistema.analisar_status_franquias()
        
        st.markdown("""
        <div class="kpi-section">
            <div class="kpi-title">📊 Relatório 1 - Franquias ATIVAS/INATIVAS</div>
        </div>
        """, unsafe_allow_html=True)
        
        ativas = len(sistema.franquias_status[sistema.franquias_status['Status'] == 'Ativo'])
        inativas = len(sistema.franquias_status[sistema.franquias_status['Status'] == 'Inativo'])
        total_franquias = len(sistema.franquias_status)
        
        col1, col2, col3, col4 = st.columns(4)
        
        with col1:
            st.markdown(f"""
            <div class="metric-card">
                <div class="icon">🏢</div>
                <h3>Total de Franquias</h3>
                <h2>{total_franquias:,}</h2>
                <small>Franquias registradas</small>
            </div>
            """, unsafe_allow_html=True)
        
        with col2:
            st.markdown(f"""
            <div class="metric-card ativo-card">
                <div class="icon">🟢</div>
                <h3>Franquias ATIVAS</h3>
                <h2>{ativas:,}</h2>
                <small>{(ativas/total_franquias*100):.1f}% do total</small>
            </div>
            """, unsafe_allow_html=True)
        
        with col3:
            st.markdown(f"""
            <div class="metric-card inativo-card">
                <div class="icon">🔴</div>
                <h3>Franquias INATIVAS</h3>
                <h2>{inativas:,}</h2>
                <small>{(inativas/total_franquias*100):.1f}% do total</small>
            </div>
            """, unsafe_allow_html=True)
        
        with col4:
            faturamento_total = sistema.faturamento_data['R$ Total'].sum()
            st.markdown(f"""
            <div class="metric-card">
                <div class="icon">💰</div>
                <h3>Faturamento Total</h3>
                <h2>R$ {faturamento_total:,.2f}</h2>
                <small>Período completo</small>
            </div>
            """, unsafe_allow_html=True)
    
    # KPIs para Contas a Receber
    if tem_contas:
        analise_contas = sistema.analisar_contas_receber()
        
        st.markdown("---")
        st.markdown("""
        <div class="kpi-section">
            <div class="kpi-title">💰 Contas a Receber</div>
        </div>
        """, unsafe_allow_html=True)
        
        col1, col2, col3, col4 = st.columns(4)
        
        with col1:
            st.markdown(f"""
            <div class="metric-card">
                <div class="icon">📄</div>
                <h3>Total de Contas</h3>
                <h2>{analise_contas['total_contas']:,}</h2>
                <small>Contas registradas</small>
            </div>
            """, unsafe_allow_html=True)
        
        with col2:
            st.markdown(f"""
            <div class="metric-card">
                <div class="icon">💵</div>
                <h3>Valor Total</h3>
                <h2>R$ {analise_contas['valor_total']:,.2f}</h2>
                <small>Valor a receber</small>
            </div>
            """, unsafe_allow_html=True)
        
        with col3:
            st.markdown(f"""
            <div class="metric-card ativo-card">
                <div class="icon">💰</div>
                <h3>Valor Pago</h3>
                <h2>R$ {analise_contas['valor_pago']:,.2f}</h2>
                <small>Valores recebidos</small>
            </div>
            """, unsafe_allow_html=True)
        
        with col4:
            st.markdown(f"""
            <div class="metric-card inativo-card">
                <div class="icon">⏳</div>
                <h3>Valor Pendente</h3>
                <h2>R$ {analise_contas['valor_pendente']:,.2f}</h2>
                <small>Valores a receber</small>
            </div>
            """, unsafe_allow_html=True)
        
        # KPIs adicionais para status de pagamento
        if not analise_contas['status_pagamento'].empty:
            st.markdown("---")
            st.markdown("""
            <div class="kpi-section">
                <div class="kpi-title">📊 Status de Pagamento</div>
            </div>
            """, unsafe_allow_html=True)
            
            col1, col2, col3 = st.columns(3)
            
            pagos = analise_contas['status_pagamento'].get('Pago', 0)
            abertos = analise_contas['status_pagamento'].get('Aberto', 0)
            total_status = pagos + abertos
            
            with col1:
                percent_pago = (pagos / total_status * 100) if total_status > 0 else 0
                st.markdown(f"""
                <div class="metric-card ativo-card">
                    <div class="icon">✅</div>
                    <h3>Contas Pagas</h3>
                    <h2>{pagos:,}</h2>
                    <small>{percent_pago:.1f}% do total</small>
                </div>
                """, unsafe_allow_html=True)
            
            with col2:
                percent_aberto = (abertos / total_status * 100) if total_status > 0 else 0
                st.markdown(f"""
                <div class="metric-card inativo-card">
                    <div class="icon">📋</div>
                    <h3>Contas Abertas</h3>
                    <h2>{abertos:,}</h2>
                    <small>{percent_aberto:.1f}% do total</small>
                </div>
                """, unsafe_allow_html=True)
            
            with col3:
                st.markdown(f"""
                <div class="metric-card">
                    <div class="icon">📊</div>
                    <h3>Total Status</h3>
                    <h2>{total_status:,}</h2>
                    <small>Contas com status</small>
                </div>
                """, unsafe_allow_html=True)
        
        # KPIs adicionais para vendedores
        if not analise_contas['contas_por_vendedor'].empty:
            st.markdown("---")
            st.markdown("""
            <div class="kpi-section">
                <div class="kpi-title">🎯 Vendedores Corrigidos</div>
            </div>
            """, unsafe_allow_html=True)
            
            col1, col2, col3 = st.columns(3)
            
            total_vendedores = len(analise_contas['contas_por_vendedor'])
            top_vendedor = analise_contas['contas_por_vendedor'].iloc[0]
            
            with col1:
                st.markdown(f"""
                <div class="metric-card">
                    <div class="icon">👥</div>
                    <h3>Total Vendedores</h3>
                    <h2>{total_vendedores}</h2>
                    <small>Vendedores ativos</small>
                </div>
                """, unsafe_allow_html=True)
            
            with col2:
                st.markdown(f"""
                <div class="metric-card">
                    <div class="icon">🏆</div>
                    <h3>Top Vendedor</h3>
                    <h2>{top_vendedor.name}</h2>
                    <small>R$ {top_vendedor['Valor Total']:,.2f}</small>
                </div>
                """, unsafe_allow_html=True)
            
            with col3:
                vencidas_top = top_vendedor['Contas Vencidas']
                total_top = top_vendedor['Quantidade']
                st.markdown(f"""
                <div class="metric-card inativo-card">
                    <div class="icon">📊</div>
                    <h3>% Vencidas Top</h3>
                    <h2>{top_vendedor['% Vencidas']:.1f}%</h2>
                    <small>{vencidas_top}/{total_top} contas</small>
                </div>
                """, unsafe_allow_html=True)
    
    # Gráficos
    st.markdown("---")
    st.markdown("""
    <div class="kpi-section">
        <div class="kpi-title">📈 Visualizações e Análises</div>
    </div>
    """, unsafe_allow_html=True)
    
    # Gráficos para Franquias
    if tem_faturamento:
        st.markdown("---")
        st.subheader("📊 Análise de Franquias")
        
        pivot_franquia, resumo_franquia = sistema.analisar_por_franquia()
        
        if pivot_franquia is not None and not pivot_franquia.empty:
            col1, col2 = st.columns(2)
            
            with col1:
                # Gráfico de pizza - Status das Franquias
                ativas_count = len(sistema.franquias_status[sistema.franquias_status['Status'] == 'Ativo'])
                inativas_count = len(sistema.franquias_status[sistema.franquias_status['Status'] == 'Inativo'])
                
                fig_status = go.Figure(data=[go.Pie(
                    labels=['ATIVAS', 'INATIVAS'],
                    values=[ativas_count, inativas_count],
                    hole=0.4,
                    marker_colors=['#28a745', '#dc3545'],
                    textinfo='label+percent+value',
                    textposition='auto'
                )])
                
                fig_status.update_layout(
                    title="Status das Franquias",
                    font=dict(size=14),
                    height=400
                )
                
                st.plotly_chart(fig_status, use_container_width=True)
            
            with col2:
                # Gráfico de barras - Top franquias por faturamento
                top_franquias = pivot_franquia.head(10)
                
                fig_franquias = go.Figure(data=[
                    go.Bar(
                        x=top_franquias['Franquias'],
                        y=top_franquias['Total Faturado'],
                        marker_color='lightblue',
                        text=top_franquias['Total Faturado'].apply(lambda x: f"R$ {x:,.0f}"),
                        textposition='auto'
                    )
                ])
                
                fig_franquias.update_layout(
                    title="Top 10 Franquias - Faturamento Total",
                    xaxis_title="Franquia",
                    yaxis_title="Faturamento (R$)",
                    height=400,
                    xaxis_tickangle=-45
                )
                
                st.plotly_chart(fig_franquias, use_container_width=True)
    
    # Gráficos para Contas a Receber
    if tem_contas:
        st.markdown("---")
        st.subheader("💰 Análise de Contas a Receber")
        
        analise_contas = sistema.analisar_contas_receber()
        
        col1, col2 = st.columns(2)
        
        with col2:
            # Gráfico de pizza - Status de Pagamento
            if not analise_contas['status_pagamento'].empty:
                fig_pagamento = go.Figure(data=[go.Pie(
                    labels=analise_contas['status_pagamento'].index,
                    values=analise_contas['status_pagamento'].values,
                    hole=0.4,
                    marker_colors=['#28a745', '#dc3545', '#6c757d'],
                    textinfo='label+percent+value',
                    textposition='auto'
                )])
                
                fig_pagamento.update_layout(
                    title="Status de Pagamento das Contas",
                    font=dict(size=14),
                    height=400
                )
                
                st.plotly_chart(fig_pagamento, use_container_width=True)
        
        # Gráficos adicionais para pagamentos
        st.markdown("---")
        st.subheader("💰 Análise de Pagamentos por Vendedor")
        
        col3, col4 = st.columns(2)
        
        with col3:
            # Gráfico de barras - Top vendedores por valor pago
            if not analise_contas['contas_por_vendedor'].empty:
                top_vendedores_pagos = analise_contas['contas_por_vendedor'].head(10)
                
                fig_pagos = go.Figure(data=[
                    go.Bar(
                        x=top_vendedores_pagos.index,
                        y=top_vendedores_pagos['Valor Pago'],
                        marker_color='#28a745',
                        text=top_vendedores_pagos['Valor Pago'].apply(lambda x: f"R$ {x:,.0f}"),
                        textposition='auto'
                    )
                ])
                
                fig_pagos.update_layout(
                    title="Top 10 Vendedores - Valor Pago",
                    xaxis_title="Vendedor",
                    yaxis_title="Valor Pago (R$)",
                    height=400,
                    xaxis_tickangle=-45
                )
                
                st.plotly_chart(fig_pagos, use_container_width=True)
        
        with col4:
            # Gráfico de barras - Top vendedores por valor pendente
            if not analise_contas['contas_por_vendedor'].empty:
                top_vendedores_pendentes = analise_contas['contas_por_vendedor'].head(10)
                
                fig_pendentes = go.Figure(data=[
                    go.Bar(
                        x=top_vendedores_pendentes.index,
                        y=top_vendedores_pendentes['Valor Pendente'],
                        marker_color='#dc3545',
                        text=top_vendedores_pendentes['Valor Pendente'].apply(lambda x: f"R$ {x:,.0f}"),
                        textposition='auto'
                    )
                ])
                
                fig_pendentes.update_layout(
                    title="Top 10 Vendedores - Valor Pendente",
                    xaxis_title="Vendedor",
                    yaxis_title="Valor Pendente (R$)",
                    height=400,
                    xaxis_tickangle=-45
                )
                
                st.plotly_chart(fig_pendentes, use_container_width=True)

def main():
    """Função principal do sistema unificado final"""
    # Inicializar sistema na sessão
    if 'sistema' not in st.session_state:
        st.session_state.sistema = SistemaUnificadoFinal()
    
    sistema = st.session_state.sistema
    
    # Sidebar
    st.sidebar.markdown("## 📁 Upload de Arquivos")
    
    # Upload do arquivo de faturamento
    faturamento_file = st.sidebar.file_uploader(
        "📊 Relatório 1 - Faturamento Franquias",
        type=['xlsx'],
        key="faturamento_upload"
    )
    
    # Upload do arquivo de contas a receber
    contas_file = st.sidebar.file_uploader(
        "💰 Relatório 2 - Contas a Receber",
        type=['xlsx'],
        key="contas_upload"
    )
    
    # Upload do arquivo de clientes (opcional)
    clientes_file = st.sidebar.file_uploader(
        "👥 Dados de Clientes (Opcional)",
        type=['xlsx'],
        key="clientes_upload"
    )
    
    # Botão para processar dados
    if st.sidebar.button("🔄 Processar Dados", type="primary"):
        with st.spinner("Processando dados..."):
            success_count = 0
            
            # Processar faturamento
            if faturamento_file is not None:
                temp_path_faturamento = "temp_faturamento.xlsx"
                with open(temp_path_faturamento, "wb") as f:
                    f.write(faturamento_file.getbuffer())
                
                if sistema.carregar_faturamento(temp_path_faturamento):
                    success_count += 1
                    os.remove(temp_path_faturamento)
            
            # Processar contas a receber
            if contas_file is not None:
                temp_path_contas = "temp_contas.xlsx"
                with open(temp_path_contas, "wb") as f:
                    f.write(contas_file.getbuffer())
                
                if sistema.carregar_contas_receber(temp_path_contas):
                    success_count += 1
                    os.remove(temp_path_contas)
            
            # Processar clientes
            if clientes_file is not None:
                temp_path_clientes = "temp_clientes.xlsx"
                with open(temp_path_clientes, "wb") as f:
                    f.write(clientes_file.getbuffer())
                
                if sistema.carregar_clientes(temp_path_clientes):
                    success_count += 1
                    os.remove(temp_path_clientes)
            
            if success_count > 0:
                st.sidebar.success(f"✅ {success_count} arquivo(s) processado(s) com sucesso!")
                st.rerun()
            else:
                st.sidebar.error("❌ Nenhum arquivo válido para processar.")
    
    # Configurações de visualização
    st.sidebar.markdown("---")
    st.sidebar.markdown("## 👁️ Configurações de Visualização")
    
    # Seleção de relatório para exibir
    relatorio_selecionado = st.sidebar.selectbox(
        "Selecione o relatório para visualizar:",
        ["📊 Ambos os Relatórios", "🏢 Apenas Franquias", "💰 Apenas Contas a Receber"]
    )
    
    # Filtros
    st.sidebar.markdown("---")
    st.sidebar.markdown("## 🔍 Filtros")
    
    # Filtro para franquias
    if sistema.faturamento_data is not None:
        st.sidebar.markdown("**📊 Filtros - Franquias**")
        
        # Filtro por status
        status_options = ["Todos", "Ativos", "Inativos"]
        status_filtro = st.sidebar.selectbox("Status da Franquia:", status_options)
        
        # Aplicar filtros
        if status_filtro != "Todos" and sistema.franquias_status is not None:
            if status_filtro == "Ativos":
                sistema.franquias_status = sistema.franquias_status[sistema.franquias_status['Status'] == 'Ativo']
            elif status_filtro == "Inativos":
                sistema.franquias_status = sistema.franquias_status[sistema.franquias_status['Status'] == 'Inativo']
    
    # Filtro para contas a receber
    if sistema.contas_receber_data is not None:
        st.sidebar.markdown("**💰 Filtros - Contas a Receber**")
        
        # Filtro por status de pagamento
        if 'Status' in sistema.contas_receber_data.columns:
            status_pagamento_options = ["Todos"] + list(sistema.contas_receber_data['Status'].unique())
            status_pagamento_filtro = st.sidebar.selectbox("Status de Pagamento:", status_pagamento_options)
            
            # Aplicar filtro
            if status_pagamento_filtro != "Todos":
                sistema.contas_receber_data = sistema.contas_receber_data[
                    sistema.contas_receber_data['Status'] == status_pagamento_filtro
                ]
    
    # Exportação
    st.sidebar.markdown("---")
    st.sidebar.markdown("## 📥 Exportação")
    
    if st.sidebar.button("📊 Gerar Relatório Completo"):
        with st.spinner("Gerando relatório..."):
            try:
                nome_relatorio = sistema.gerar_relatorio_completo()
                st.sidebar.success(f"✅ Relatório gerado: {nome_relatorio}")
                
                # Oferecer download
                with open(nome_relatorio, "rb") as file:
                    st.sidebar.download_button(
                        label="📥 Baixar Relatório Excel",
                        data=file.read(),
                        file_name=nome_relatorio,
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                    )
            except Exception as e:
                st.sidebar.error(f"❌ Erro ao gerar relatório: {str(e)}")
    
    # Mostrar dashboard baseado na seleção
    if relatorio_selecionado == "📊 Ambos os Relatórios":
        mostrar_dashboard(sistema)
    elif relatorio_selecionado == "🏢 Apenas Franquias":
        # Mostrar apenas dashboard de franquias
        if sistema.faturamento_data is not None:
            # Temporariamente esconder contas para mostrar apenas franquias
            temp_contas = sistema.contas_receber_data
            sistema.contas_receber_data = None
            mostrar_dashboard(sistema)
            sistema.contas_receber_data = temp_contas
        else:
            st.warning("⚠️ Nenhum dado de faturamento carregado.")
    elif relatorio_selecionado == "💰 Apenas Contas a Receber":
        # Mostrar apenas dashboard de contas
        if sistema.contas_receber_data is not None:
            # Temporariamente esconder faturamento para mostrar apenas contas
            temp_faturamento = sistema.faturamento_data
            sistema.faturamento_data = None
            mostrar_dashboard(sistema)
            sistema.faturamento_data = temp_faturamento
        else:
            st.warning("⚠️ Nenhum dado de contas a receber carregado.")
    
    # Tabelas de dados detalhados
    if sistema.faturamento_data is not None or sistema.contas_receber_data is not None:
        st.markdown("---")
        st.markdown("""
        <div class="kpi-section">
            <div class="kpi-title">📋 Dados Detalhados</div>
        </div>
        """, unsafe_allow_html=True)
        
        # Abas para diferentes visualizações
        tab1, tab2, tab3 = st.tabs(["📊 Franquias", "💰 Contas a Receber", "👥 Clientes"])
        
        with tab1:
            if sistema.franquias_status is not None:
                st.markdown("#### 📊 Status das Franquias")
                st.dataframe(
                    sistema.franquias_status,
                    use_container_width=True,
                    hide_index=True
                )
            
            # Tabela detalhada por franquia
            if sistema.faturamento_data is not None:
                pivot_franquia, resumo_franquia = sistema.analisar_por_franquia()
                if resumo_franquia is not None and not resumo_franquia.empty:
                    st.markdown("#### 📈 Resumo por Franquia")
                    st.dataframe(
                        resumo_franquia,
                        use_container_width=True,
                        hide_index=True
                    )
        
        with tab2:
            if sistema.contas_receber_data is not None:
                st.markdown("#### 💰 Contas a Receber")
                
                # Análise para exibição
                analise_contas = sistema.analisar_contas_receber()
                
                # Tabela de contas
                st.dataframe(
                    sistema.contas_receber_data,
                    use_container_width=True,
                    hide_index=True
                )
                
                # Tabela por vendedor
                if not analise_contas['contas_por_vendedor'].empty:
                    st.markdown("#### 🎯 Análise por Vendedor")
                    st.dataframe(
                        analise_contas['contas_por_vendedor'],
                        use_container_width=True
                    )
        
        with tab3:
            if sistema.clientes_data is not None:
                st.markdown("#### 👥 Dados dos Clientes")
                st.dataframe(
                    sistema.clientes_data,
                    use_container_width=True,
                    hide_index=True
                )
            else:
                st.info("📋 Nenhum dado de clientes carregado. Faça upload do arquivo de clientes para visualizar.")

if __name__ == "__main__":
    main()
